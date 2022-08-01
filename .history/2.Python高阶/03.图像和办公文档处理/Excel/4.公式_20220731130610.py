import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '公式'
sheet['A1'].value = 100
sheet['A2'].value = 200
sheet['A3'].value = 300
sheet['A4'].value = '=SUM(A1:A3)'
print(sheet['A4'].value)  # 结果是公式SUM(A1:A3)
wb.save(r'./example01.xlsx')

# 调用 load_workbook()时带有 data_only=True，则只会加载数据，不会加载公式
wb = openpyxl.load_workbook(r'./example01.xlsx', data_only=True)
sheet = wb.active
print(sheet['A1'].value)  # 结果是600
