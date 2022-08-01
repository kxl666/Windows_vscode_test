import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'].value = '行高'
sheet['B1'].value = '列宽'

# 1.设置行高和列宽
sheet.column_dimensions['A'].width = 50
sheet.column_dimensions['B'].width = 70
sheet.row_dimensions[1].height = 30

# 2.合并和拆分单元格
# sheet.merge_cells('A1:B1')
# sheet.unmerge_cells('A1:B1')
wb.save(r'./File/example02.xlsx')
