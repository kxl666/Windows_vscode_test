import openpyxl

# 使用openpyxl模块创建新的空Excel文件
wb = openpyxl.Workbook()
# 获取活动的工作表
sheet = wb.active
# 修改工作表名称
sheet.title = 'Now Sheet'

# 1.创建和删除工作表
# 创建新的工作表
wb.create_sheet('New Sheet')
# 获取所有工作表的名称
print(wb.sheetnames)
# 通过索引方式插入工作表名称
wb.create_sheet(index=1, title='Change Sheet')
print(wb.sheetnames)

# 删除工作表
