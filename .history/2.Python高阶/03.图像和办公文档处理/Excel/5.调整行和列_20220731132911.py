import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'].value = '行高'
sheet['B1'].value = '列宽'
sheet['C1'].value = '边框'
sheet['A2'].value = '行高'
sheet['A3'].value = '行高'
sheet['A4'].value = '列宽'
sheet['A5'].value = '边框'
sheet['A6'].value = '行高'

# 1.设置行高和列宽
sheet.column_dimensions['A'].width = 50
sheet.column_dimensions['B'].width = 70
sheet.row_dimensions[1].height = 30

# 2.合并和拆分单元格
# sheet.merge_cells('A1:B1')
# sheet.unmerge_cells('A1:B1')

# 3.冻结窗格
# sheet.freeze_panes = 'A2'

# 4.设置边框
sheet['A1'].border = openpyxl.styles.Border(left=openpyxl.styles.Side(
    style='thin'))
sheet['A2'].border = openpyxl.styles.Border(top=openpyxl.styles.Side(
    style='thin'))

# 5.设置背景色
sheet['A1'].fill = openpyxl.styles.PatternFill(patternType='solid',
                                               fgColor='FF0000')

# 6.图表

wb.save(r'./File/example02.xlsx')
