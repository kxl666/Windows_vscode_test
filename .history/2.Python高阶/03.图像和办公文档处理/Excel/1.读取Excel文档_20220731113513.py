import openpyxl

# 使用openpyxl模块打开Excel文件
wb = openpyxl.load_workbook(r'./example.xlsx')

# 1.获取工作表
# 获取工作簿中的所有工作表
print(wb.get_sheet_names())
# 获取工作簿中的第一个工作表
sheet = wb.get_sheet_by_name('Sheet1')
# 修改工作表名称
sheet.title = 'New Sheet'
print(wb.get_sheet_names())

# 2.获取单元格
# 获取单元格的值
print(sheet['A1'].value)
# 获取单元格的坐标
print(sheet['A1'].coordinate)
# 获取单元格的行号和列号
print(sheet['A1'].row, sheet['A1'].column)
# Cell对象
print(sheet.cell(row=1, column=2).value)
# 循环遍历单元格
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        print(sheet.cell(row=row, column=column).value, end='\t')
    print()

# 3.获取行和列
# 获取行对象
print(sheet.row_dimensions[1].style)
# 获取列对象
print(sheet.column_dimensions['B'])
# 获取行的高度
print(sheet.row_dimensions[1].height)
